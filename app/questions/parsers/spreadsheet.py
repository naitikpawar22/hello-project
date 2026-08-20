<<<<<<< HEAD
import re
=======
>>>>>>> 00cdc5ce5c2c164af42ff31e6595073d105d2b0b
from openpyxl import load_workbook
from .common import make_question

def parse(path):
<<<<<<< HEAD
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active; rows = ws.iter_rows(values_only=True)
    raw_headers = next(rows, ())
    headers = [re.sub(r'[\s\.\-]+', '_', str(x or "").strip().lower()) for x in raw_headers]
    records = []
    for vals in rows:
        if not vals: continue
        row = {headers[i]: str(vals[i] if vals[i] is not None else "").strip() for i in range(min(len(headers), len(vals)))}
        
        q = row.get("question") or row.get("question_text")
        if not q: continue
        
        opts = []
        for opt_key in ("option_a", "option_b", "option_c", "option_d", "option_1", "option_2", "option_3", "option_4", "a", "b", "c", "d"):
            val = row.get(opt_key)
            if val: opts.append(val)
            
        answer = row.get("correct_answer") or row.get("correct") or row.get("answer")
        topic = row.get("category") or row.get("topic") or ""
        typ = row.get("type") or row.get("question_type") or "MCQ"
        marks = row.get("marks", 1)
        neg = row.get("negative_marks", 0)
        expl = row.get("explanation") or row.get("solution") or ""
        
        records.append(make_question(q, opts, answer, type=typ, marks=marks, negative_marks=neg, topic=topic, explanation=expl))
=======
    wb=load_workbook(path,read_only=True,data_only=True)
    ws=wb.active; rows=ws.iter_rows(values_only=True)
    headers=[str(x or "").strip().lower() for x in next(rows,())]
    records=[]
    for vals in rows:
        row={headers[i]:vals[i] for i in range(min(len(headers),len(vals)))}
        q=row.get("question") or row.get("question_text")
        if not q: continue
        options=[row.get(k) for k in ("option_a","option_b","option_c","option_d") if row.get(k) not in (None,"")]
        records.append(make_question(q,options,row.get("answer") or row.get("correct_answer"),type=row.get("type") or row.get("question_type") or "MCQ",marks=row.get("marks",1),negative_marks=row.get("negative_marks",0),topic=row.get("topic") or row.get("category") or "",explanation=row.get("explanation") or row.get("solution") or ""))
>>>>>>> 00cdc5ce5c2c164af42ff31e6595073d105d2b0b
    return records
